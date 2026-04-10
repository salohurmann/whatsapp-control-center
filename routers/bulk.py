import csv
import io
import json
import re
import unicodedata

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from dependencies import get_client_id, get_operator_name, require_api_key
from services import audit, bulk_manager, client_manager, suppression


router = APIRouter(dependencies=[Depends(require_api_key)])

MAX_CONTACTS = 10000


@router.post("/send", summary="Criar campanha de disparo em massa")
async def bulk_send(
    file: UploadFile = File(..., description="CSV ou Excel (.xlsx) com coluna 'telefone'"),
    message: str = Form(..., description="Mensagem a enviar para todos"),
    delay_seconds: float = Form(1.0, description="Intervalo entre envios em segundos"),
    pilot_size: int = Form(0, description="Se maior que zero, cria uma campanha piloto limitada aos primeiros contatos."),
    client_id: str = Depends(get_client_id),
    operator: str = Depends(get_operator_name),
):
    if delay_seconds < 0.5:
        raise HTTPException(status_code=400, detail="delay_seconds minimo e 0.5 para evitar bloqueio e rate limit.")

    client = client_manager.resolve_client(client_id)
    if not client.get("simulation_mode") and not (client.get("access_token") and client.get("phone_number_id")):
        raise HTTPException(status_code=400, detail="Configure ACCESS_TOKEN e PHONE_NUMBER_ID do cliente antes de disparar.")

    content = await file.read()
    filename = (file.filename or "").lower()
    contacts = _parse_excel(content) if filename.endswith((".xlsx", ".xls")) else _parse_csv(content)

    if not contacts:
        raise HTTPException(status_code=422, detail="Nenhum numero valido encontrado no arquivo.")
    if len(contacts) > MAX_CONTACTS:
        contacts = contacts[:MAX_CONTACTS]
    if pilot_size > 0:
        contacts = contacts[:pilot_size]

    job = bulk_manager.create_job(client_id=client_id, message=message, delay_seconds=delay_seconds, contacts=contacts)
    audit.record_event(
        client_id=client_id,
        event_type="job_created",
        entity_type="job",
        entity_id=job["id"],
        operator=operator,
        details={"total": job["total"], "pilot_size": pilot_size},
    )
    return {
        "job_id": job["id"],
        "client_id": client_id,
        "status": job["status"],
        "total_contacts": job["total"],
        "deduplicated": job["deduplicated"],
        "message": "Campanha criada e adicionada a fila.",
        "acompanhe": f"/bulk/status/{job['id']}",
        "relatorio": f"/bulk/report/{job['id']}",
    }


@router.post("/send-direct", summary="Criar campanha em massa a partir de numeros colados")
async def bulk_send_direct(
    phones_text: str = Form(..., description="Lista de numeros, um por linha ou separados por virgula"),
    message: str = Form(..., description="Mensagem a enviar para todos"),
    delay_seconds: float = Form(1.0, description="Intervalo entre envios em segundos"),
    pilot_size: int = Form(0, description="Se maior que zero, cria uma campanha piloto limitada aos primeiros contatos."),
    client_id: str = Depends(get_client_id),
    operator: str = Depends(get_operator_name),
):
    if delay_seconds < 0.5:
        raise HTTPException(status_code=400, detail="delay_seconds minimo e 0.5 para evitar bloqueio e rate limit.")

    client = client_manager.resolve_client(client_id)
    if not client.get("simulation_mode") and not (client.get("access_token") and client.get("phone_number_id")):
        raise HTTPException(status_code=400, detail="Configure ACCESS_TOKEN e PHONE_NUMBER_ID do cliente antes de disparar.")

    contacts = _parse_direct_phones(phones_text)

    if not contacts:
        raise HTTPException(status_code=422, detail="Nenhum numero valido encontrado no texto informado.")
    if len(contacts) > MAX_CONTACTS:
        contacts = contacts[:MAX_CONTACTS]
    if pilot_size > 0:
        contacts = contacts[:pilot_size]

    job = bulk_manager.create_job(client_id=client_id, message=message, delay_seconds=delay_seconds, contacts=contacts)
    audit.record_event(
        client_id=client_id,
        event_type="job_created_direct",
        entity_type="job",
        entity_id=job["id"],
        operator=operator,
        details={"total": job["total"], "pilot_size": pilot_size},
    )
    return {
        "job_id": job["id"],
        "client_id": client_id,
        "status": job["status"],
        "total_contacts": job["total"],
        "deduplicated": job["deduplicated"],
        "message": "Campanha direta criada e adicionada a fila.",
        "acompanhe": f"/bulk/status/{job['id']}",
        "relatorio": f"/bulk/report/{job['id']}",
    }


@router.post("/send-template", summary="Criar campanha em massa por template aprovado")
async def bulk_send_template(
    file: UploadFile = File(..., description="CSV ou Excel (.xlsx) com coluna 'telefone'"),
    template_name: str = Form(..., description="Nome do template aprovado na Meta"),
    language_code: str = Form("pt_BR", description="Idioma do template"),
    components_json: str = Form("[]", description="JSON com os componentes do template"),
    delay_seconds: float = Form(1.0, description="Intervalo entre envios em segundos"),
    pilot_size: int = Form(0, description="Se maior que zero, cria uma campanha piloto limitada aos primeiros contatos."),
    client_id: str = Depends(get_client_id),
    operator: str = Depends(get_operator_name),
):
    if delay_seconds < 0.5:
        raise HTTPException(status_code=400, detail="delay_seconds minimo e 0.5 para evitar bloqueio e rate limit.")

    client = client_manager.resolve_client(client_id)
    if not client.get("simulation_mode") and not (client.get("access_token") and client.get("phone_number_id")):
        raise HTTPException(status_code=400, detail="Configure ACCESS_TOKEN e PHONE_NUMBER_ID do cliente antes de disparar.")

    try:
        template_components = json.loads(components_json or "[]")
        if not isinstance(template_components, list):
            raise ValueError("components_json deve ser uma lista JSON.")
    except Exception as exc:
        raise HTTPException(status_code=422, detail="components_json invalido.") from exc

    content = await file.read()
    filename = (file.filename or "").lower()
    contacts = _parse_excel(content) if filename.endswith((".xlsx", ".xls")) else _parse_csv(content)

    if not contacts:
        raise HTTPException(status_code=422, detail="Nenhum numero valido encontrado no arquivo.")
    if len(contacts) > MAX_CONTACTS:
        contacts = contacts[:MAX_CONTACTS]
    if pilot_size > 0:
        contacts = contacts[:pilot_size]

    job = bulk_manager.create_job(
        client_id=client_id,
        message=f"[template] {template_name}",
        delay_seconds=delay_seconds,
        contacts=contacts,
        send_mode="template",
        template_name=template_name,
        language_code=language_code,
        template_components=template_components,
    )
    audit.record_event(
        client_id=client_id,
        event_type="job_created_template",
        entity_type="job",
        entity_id=job["id"],
        operator=operator,
        details={"total": job["total"], "pilot_size": pilot_size, "template_name": template_name},
    )
    return {
        "job_id": job["id"],
        "client_id": client_id,
        "status": job["status"],
        "total_contacts": job["total"],
        "template_name": template_name,
        "message": "Campanha de template criada e adicionada a fila.",
    }


def _find_col(fieldnames: list[str], accepted: list[str]):
    accepted_normalized = {item.strip().lower() for item in accepted}
    return next((f for f in fieldnames if _normalize_header(f) in accepted_normalized), None)


def _find_col_idx(headers: list[str], accepted: list[str]):
    accepted_normalized = {item.strip().lower() for item in accepted}
    for i, header in enumerate(headers):
        if _normalize_header(header) in accepted_normalized:
            return i
    return None


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    safe = "".join(ch.lower() if ch.isalnum() else "_" for ch in ascii_only)
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe.strip("_")


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    first_line = text.splitlines()[0] if text.splitlines() else ""
    sep = ";" if ";" in first_line else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    phone_col = _find_col(reader.fieldnames or [], ["telefone", "phone", "numero", "número", "cel", "celular"])
    name_col = _find_col(reader.fieldnames or [], ["nome", "name"])

    if not phone_col:
        raise HTTPException(status_code=422, detail=f"Coluna 'telefone' nao encontrada. Colunas: {reader.fieldnames}")

    contacts: list[dict[str, str]] = []
    for row in reader:
        phone = "".join(filter(str.isdigit, row.get(phone_col, "")))
        if 10 <= len(phone) <= 15:
            contact = {"phone": phone, "name": row.get(name_col, "") if name_col else ""}
            for key, value in row.items():
                normalized_key = _normalize_header(key)
                if normalized_key:
                    contact[normalized_key] = str(value or "").strip()
            contacts.append(contact)
        if len(contacts) >= MAX_CONTACTS:
            break
    return contacts


def _parse_direct_phones(phones_text: str) -> list[dict[str, str]]:
    raw_items = re.split(r"[\n,;]+", phones_text or "")
    contacts: list[dict[str, str]] = []
    for raw in raw_items:
        phone = "".join(filter(str.isdigit, raw))
        if 10 <= len(phone) <= 15:
            contacts.append({"phone": phone, "name": ""})
        if len(contacts) >= MAX_CONTACTS:
            break
    return contacts


def _parse_excel(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl nao instalado.") from exc

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(item).strip() if item is not None else "" for item in next(rows_iter, [])]
    normalized_headers = [_normalize_header(item) for item in headers]
    phone_idx = _find_col_idx(headers, ["telefone", "phone", "numero", "número", "cel", "celular"])
    name_idx = _find_col_idx(headers, ["nome", "name"])

    if phone_idx is None:
        raise HTTPException(status_code=422, detail=f"Coluna 'telefone' nao encontrada. Colunas: {headers}")

    contacts: list[dict[str, str]] = []
    for row in rows_iter:
        raw_phone = row[phone_idx] if phone_idx < len(row) else None
        phone = "".join(filter(str.isdigit, str(raw_phone or "")))
        if 10 <= len(phone) <= 15:
            name = str(row[name_idx] or "").strip() if name_idx is not None and name_idx < len(row) else ""
            contact = {"phone": phone, "name": name}
            for index, normalized_header in enumerate(normalized_headers):
                if normalized_header and index < len(row):
                    contact[normalized_header] = str(row[index] or "").strip()
            contacts.append(contact)
        if len(contacts) >= MAX_CONTACTS:
            break
    wb.close()
    return contacts


@router.get("/status/{job_id}", summary="Acompanhar progresso da campanha")
async def bulk_status(job_id: str, client_id: str = Depends(get_client_id)):
    job = bulk_manager.get_job_status(job_id, client_id=client_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job nao encontrado.")
    return {
        "job_id": job["id"],
        "client_id": job["client_id"],
        "status": job["status"],
        "progresso": job["progress_text"],
        "enviados": job["sent"],
        "erros": job["failed"],
        "falhas_temporarias": job["temporary_failures"],
        "deduplicados": job["deduplicated"],
        "criado_em": job["created_at"],
        "iniciado_em": job["started_at"],
        "finalizado_em": job["finished_at"],
        "last_error": job["last_error"],
        "delivery": job["delivery"],
        "recent_contacts": job["recent_contacts"],
    }


@router.get("/report/{job_id}", summary="Baixar relatorio CSV da campanha")
async def bulk_report(job_id: str, client_id: str = Depends(get_client_id)):
    output = bulk_manager.get_report_csv(job_id, client_id=client_id)
    if output is None:
        raise HTTPException(status_code=404, detail="Job nao encontrado.")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{job_id}.csv"},
    )


@router.get("/jobs", summary="Listar campanhas")
async def list_jobs(
    client_id: str = Depends(get_client_id),
    status: str | None = Query(default=None),
    q: str | None = Query(default=None),
):
    jobs = bulk_manager.list_jobs(client_id=client_id)
    if status:
        jobs = [job for job in jobs if job["status"] == status]
    if q:
        q_lower = q.lower()
        jobs = [job for job in jobs if q_lower in job["id"].lower()]
    return [
        {
            "job_id": job["id"],
            "client_id": job["client_id"],
            "status": job["status"],
            "total": job["total"],
            "enviados": job["sent"],
            "erros": job["failed"],
            "falhas_temporarias": job["temporary_failures"],
            "deduplicados": job["deduplicated"],
            "criado_em": job["created_at"],
            "iniciado_em": job["started_at"],
            "finalizado_em": job["finished_at"],
            "updated_at": job["updated_at"],
            "last_error": job["last_error"],
        }
        for job in jobs
    ]


@router.get("/dashboard", summary="Resumo operacional dos disparos")
async def dashboard(client_id: str = Depends(get_client_id)):
    return bulk_manager.get_dashboard_stats(client_id=client_id)


@router.post("/pause/{job_id}", summary="Pausar campanha")
async def pause_job(job_id: str, client_id: str = Depends(get_client_id)):
    job = bulk_manager.request_pause(job_id, client_id=client_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job nao encontrado.")
    audit.record_event(client_id=client_id, event_type="job_pause_requested", entity_type="job", entity_id=job_id)
    return {"job_id": job_id, "status": job["status"]}


@router.post("/resume/{job_id}", summary="Retomar campanha")
async def resume_job(job_id: str, client_id: str = Depends(get_client_id)):
    job = bulk_manager.request_resume(job_id, client_id=client_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job nao encontrado.")
    audit.record_event(client_id=client_id, event_type="job_resumed", entity_type="job", entity_id=job_id)
    return {"job_id": job_id, "status": job["status"]}


@router.post("/cancel/{job_id}", summary="Cancelar campanha")
async def cancel_job(job_id: str, client_id: str = Depends(get_client_id)):
    job = bulk_manager.request_cancel(job_id, client_id=client_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job nao encontrado.")
    audit.record_event(client_id=client_id, event_type="job_cancel_requested", entity_type="job", entity_id=job_id)
    return {"job_id": job_id, "status": job["status"]}


@router.post("/retry-temporary/{job_id}", summary="Reenfileirar falhas temporarias")
async def retry_temporary(job_id: str, client_id: str = Depends(get_client_id)):
    if not bulk_manager.get_job(job_id, client_id=client_id):
        raise HTTPException(status_code=404, detail="Job nao encontrado.")
    retried = bulk_manager.reset_temporary_failures(job_id, client_id=client_id)
    audit.record_event(client_id=client_id, event_type="job_retry_temporary", entity_type="job", entity_id=job_id, details={"count": retried})
    return {"job_id": job_id, "reenfileirados": retried, "status": "queued"}


@router.get("/suppression", summary="Listar supressoes")
async def list_suppression(client_id: str = Depends(get_client_id)):
    return suppression.list_phones(client_id=client_id)


@router.post("/suppression", summary="Adicionar numero a supressao")
async def add_suppression(
    phone: str = Form(...),
    reason: str = Form(""),
    client_id: str = Depends(get_client_id),
):
    try:
        result = suppression.add_phone(phone, client_id=client_id, reason=reason, source="manual")
        audit.record_event(client_id=client_id, event_type="suppression_added", entity_type="suppression", entity_id=result["phone"], details=result)
        return result
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@router.delete("/suppression/{phone}", summary="Remover numero da supressao")
async def remove_suppression(phone: str, client_id: str = Depends(get_client_id)):
    removed = suppression.remove_phone(phone, client_id=client_id)
    if not removed:
        raise HTTPException(status_code=404, detail="Numero nao encontrado na supressao.")
    audit.record_event(client_id=client_id, event_type="suppression_removed", entity_type="suppression", entity_id=phone)
    return {"removed": True, "phone": phone}


@router.get("/contacts", summary="Listar contatos processados")
async def list_contacts(
    client_id: str = Depends(get_client_id),
    job_id: str | None = Query(default=None),
    status: str | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
):
    return bulk_manager.list_contacts(client_id=client_id, job_id=job_id, status=status, q=q, limit=limit)
